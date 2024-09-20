# consider using pyexcel library too.

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pdb

os.chdir("C:/Users/knkas/Desktop/NLP_example")

data = {'Name': ['Alice', 'Bob', 'Charlie'],
        'Age': [25, 30, 35],
        'City': ['New York', 'Los Angeles', 'Chicago']}
df = pd.DataFrame(data)

file_path = 'my_data.xlsx'
sheet_name = "SheetX"
sheet_name2 = "SheetY"

with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:  # Use this if you want to append another sheet to the existing file.
    df.to_excel(writer, sheet_name=sheet_name, index=False, startcol=4, startrow=4)
    df.to_excel(writer, sheet_name=sheet_name2, index=False, startcol=6, startrow=6)

# Insert values to a specific cell in an existing excel file.
wb = load_workbook(file_path) # load the existing file.
ws = wb['SheetX']
start_cell = "A1"

# Replace column names in the Excel sheet
for c_idx, new_col_name in enumerate(df.columns):
    cell = ws.cell(row=int(start_cell[1:]), column=ord(start_cell[0]) - ord('A') + c_idx + 1)
    cell.value = new_col_name

# Write values underneath the columns
for r_idx, row in df.iterrows():
    for c_idx, value in enumerate(row):
        cell = ws.cell(row=int(start_cell[1:]) + r_idx + 1, column=ord(start_cell[0]) - ord('A') + c_idx + 1)
        cell.value = value
        
        
wb.save(file_path)

